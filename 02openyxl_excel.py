# pip install openpyxl
import openpyxl
import os
# os.chdir 是 python 切換到電腦指定路徑的方法
os.chdir(r"D:\python_effective")
# 開啓 excel 檔案，存入 workbook 這個變數中
workbook = openpyxl.load_workbook('stock_price_data.xlsx')
# 從 workbook 中開啓一個名爲 2330 的工作表，存入 sheet 變數
sheet = workbook.get_sheet_by_name('2330')

########################################
# # 讀取 row(行英) 為 1, column(列數) 為 1 的儲存格
# result = sheet.cell(row=1, column=1).value
# # 日期
# print(result)

#######################################
# # (B3 - B2) / B2
# result = (sheet.cell(row=3, column=2).value - sheet.cell(row=2, column=2).value) / sheet.cell(row=2, column=2).value

# print(result)
## 0.0
################################################
for i in range(3, 100):
    # 算出每一個 row 的報酬率
    result = (sheet.cell(row=i, column=2).value - sheet.cell(row=i-1, column=2).value) / sheet.cell(row=i-1, column=2).value
    # 印出報酬率
    print(str(i)+"==>"+str(result))
    print("%d ==>%3.4f" %(i,result))
    sheet.cell(row=i,column=3).value=result
workbook.save('stock_price_data.xlsx')

