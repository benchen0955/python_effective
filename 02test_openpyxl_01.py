from openpyxl.styles import colors
from openpyxl.styles import Font, Color

from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import Workbook
from openpyxl import load_workbook
import os
file_name='balances.xlsx'
file_path = "D:/python_effective/"
os.chdir(file_path)
print(file_path+file_name)
wooksheet_list=[]
wooksheet_name=""
ws0_name="sheet_A"
ws1_name="sheet_B"
ws2_name="sheet_C" 

if ( os.path.isfile(file_name)):
    print(file_name+" File exists!")
    wb = load_workbook(file_name)
    # print("length==>"+len(wb))
    print(wb)
    wb_name_list=wb.sheetnames
    print(wb_name_list)
    
    for sheet in wb:
        print(sheet.title+" working")
        wooksheet=wb[sheet.title]
        if sheet.title == "sheet_03":
            wooksheet.cell(row=4, column=4).value="qwert"
            print(sheet.title+" done") 

    # ws0 = wb[ws0_name]
    # ws1 = wb[ws1_name]
    # ws2 = wb[ws2_name]
    # ws2.cell(row=4, column=4).value="qwert"
    # print()
    
else:
    print(file_name+" not found! Create one")
    wb = Workbook()
    wooksheet = wb.active
    wooksheet_name ="sheet_01"
    wooksheet.title = wooksheet_name
    wooksheet.sheet_properties.tabColor = "00ff00"
    wooksheet = wb[wooksheet_name]
    wooksheet_list.append(wooksheet)

    wooksheet_name ="sheet_02"
    wooksheet = wb.create_sheet(wooksheet_name)#append in end
    wooksheet.sheet_properties.tabColor = "0000ff"
    wooksheet_list.append(wooksheet)

    wooksheet_name ="sheet_03"
    wooksheet = wb.create_sheet(wooksheet_name,0)#append in first
    wooksheet.sheet_properties.tabColor = "ff0000"
    wooksheet_list.append(wooksheet)
    
    wb_name_list=wb.sheetnames
    print(wb_name_list)

    ft1 = Font(color=colors.RED, italic=True)
    ft2 = Font(color="00ff00", italic=True)

    for sheet in wb:
        print(sheet.title+" working")
        if sheet.title == "sheet_03":
            # wooksheet = wb.get_sheet_by_name(sheet.title)
            wooksheet = wb[sheet.title]       
            wooksheet['A1']=4
            wooksheet['A1'].font=ft1
            wooksheet['A1'].fill = PatternFill("solid", fgColor="0000ff")
            x=3
            y=3
            wooksheet.cell(row=y, column=x).value=23
            wooksheet.cell(row=y, column=x).font=ft2
            wooksheet.cell(row=y, column=x).fill = GradientFill(stop=("ffffff", "0000ff")) 

        if sheet.title == "sheet_01":
            # wooksheet = wb.get_sheet_by_name(sheet.title)
            wooksheet = wb[sheet.title ] 
            wooksheet['A4'] = 4
        
        print(sheet.title+" done")


    # wb_name_list[0]['A4']=
    # ws0['A4'] = 4
    # ws2['A1'] = 1
    # x=3
    # y=3
    # ws2.cell(row=y, column=x).value=23
    # ft1 = Font(color=colors.RED, italic=True)
    # ft2 = Font(color="00ff00", italic=True)
    # ws2['A1'].font=ft1
    # ws2.cell(row=y, column=x).font=ft2
    # ws2['A1'].fill = PatternFill("solid", fgColor="0000ff")
    # ws2.cell(row=y, column=x).fill = GradientFill(stop=("ffffff", "0000ff"))
    
wb.save(file_name)




